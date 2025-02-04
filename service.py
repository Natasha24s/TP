import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Color
import os
import warnings

# Suppress specific warnings
warnings.filterwarnings('ignore', category=FutureWarning)

def clean_amount(value):
    """Clean and convert monetary values to float"""
    try:
        if pd.isna(value):
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            if '(free tier)' in value:
                return 0.0
            cleaned = value.replace('US$', '').replace('$', '').replace(',', '').replace(' ', '')
            if cleaned == '-' or cleaned == '':
                return 0.0
            return float(cleaned)
        return 0.0
    except Exception as e:
        print(f"Warning: Could not convert '{value}' to float. Using 0.0 instead.")
        return 0.0

def clean_excel_data(excel_file_path):
    """Clean and standardize Excel data"""
    excel_file = pd.ExcelFile(excel_file_path)
    cleaned_sheets = {}
    
    for sheet_name in excel_file.sheet_names:
        try:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
            if df.empty or df.dropna().empty:
                print(f"Sheet '{sheet_name}' is empty - skipping")
                continue
            
            print(f"\nProcessing sheet: {sheet_name}")
            print(f"Original shape: {df.shape}")
            
            new_columns = []
            for col in df.columns:
                if isinstance(col, str):
                    if 'Service group' in col:
                        new_columns.append('Service group')
                    elif '12 months total' in col:
                        new_columns.append('Last 12 months total')
                    else:
                        new_columns.append(col)
                elif isinstance(col, datetime):
                    new_columns.append(col.strftime('%b-%y'))
                else:
                    new_columns.append(col)
            
            df.columns = new_columns
            df = df.replace('-', 0)
            
            if not df.empty:
                cleaned_sheets[sheet_name] = df
                print(f"Shape after cleaning: {df.shape}")
                print("\nColumn names after cleaning:")
                print(df.columns.tolist())
            
        except Exception as e:
            print(f"Error processing sheet '{sheet_name}': {str(e)}")
            continue
    
    return cleaned_sheets

def analyze_aws_services():
    """Analyze service categories across accounts with separate sheets for each category"""
    try:
        print("Loading data from MyPodData.xlsx...")
        excel_file_path = 'MyPodData.xlsx'
        cleaned_data = clean_excel_data(excel_file_path)
        
        if not cleaned_data:
            raise ValueError("No data found in Excel file")

        # Define service categories with their associated services
        service_categories = {
            'Storage': [
                'S3', 'EBS', 'EFS', 'Storage', 'Glacier', 'Elastic Block Store',
                'Backup', 'Transfer', 'FSx', 'Storage Gateway', 'Simple Storage Service'
            ],
            'Machine Learning & Deep Learning': [
                'SageMaker', 'Rekognition', 'Comprehend', 'ML', 'AI', 'Deep Learning',
                'Forecast', 'Textract', 'Polly', 'Transcribe', 'Bedrock', 'Lex',
                'Personalize', 'Translate', 'DeepLens', 'DeepRacer'
            ],
            'Edge': [
                'CloudFront', 'Route53', 'PerimeterProtection', 'Edge', 'CDN',
                'DNS', 'Cloud Front', 'Global Accelerator', 'Route 53'
            ],
            'Compute Services': [
                'EC2', 'ECS', 'EKS', 'Fargate', 'Lambda', 'Compute', 'Container',
                'NAT Gateway', 'Elastic Load Balancing', 'Load Balancer', 'Elastic Compute',
                'Elastic Container', 'Elastic Kubernetes', 'AWS Lambda', 'Savings Plan'
            ],
            'DB Services': [
                'RDS', 'DynamoDB', 'Elasticache', 'Aurora', 'Database', 'DDB',
                'MySQL', 'PostgreSQL', 'MariaDB', 'SQL Server', 'Neptune', 'DocumentDB',
                'Relational Database', 'NoSQL', 'Redis', 'Memcached'
            ]
        }

        # Process each service category
        dataframes = {}
        
        for category, services in service_categories.items():
            print(f"\nAnalyzing {category}...")
            analysis_data = []

            for account_name, df in cleaned_data.items():
                try:
                    category_total = 0
                    service_totals = {}

                    # First pass - find the category total
                    for idx, row in df.iterrows():
                        service_name = str(row.iloc[0])
                        if service_name == category:
                            category_total = clean_amount(row.iloc[1])
                            break

                    if category_total > 0:
                        # Second pass - find individual service totals
                        for idx, row in df.iterrows():
                            service_name = str(row.iloc[0])
                            if any(service.lower() in service_name.lower() for service in services):
                                spend = clean_amount(row.iloc[1])
                                service_totals[service_name] = spend

                        # Create row data
                        row_data = {
                            'Account': account_name,
                            f'Total {category} Spend': category_total
                        }

                        # Add service data
                        for service_name, spend in service_totals.items():
                            clean_name = service_name.replace('/', '_').replace('\\', '_')
                            percentage = (spend / category_total * 100) if category_total > 0 else 0
                            row_data[f"{clean_name} Spend"] = spend
                            row_data[f"{clean_name} %"] = percentage

                        analysis_data.append(row_data)
                        print(f"{category} total for {account_name}: ${category_total:,.2f}")

                except Exception as e:
                    print(f"Error processing {account_name} for {category}: {str(e)}")
                    continue

            # Create DataFrame for category
            if analysis_data:
                df_category = pd.DataFrame(analysis_data)

                # Add industry average row
                avg_row = {'Account': 'Industry Average'}
                for col in df_category.columns:
                    if col != 'Account':
                        avg_row[col] = df_category[col].mean()
                
                df_category = pd.concat([df_category, pd.DataFrame([avg_row])], ignore_index=True)

                # Format numeric columns
                for col in df_category.columns:
                    if 'Spend' in col:
                        df_category[col] = df_category[col].apply(lambda x: f"${float(x):,.2f}")
                    elif '%' in col:
                        df_category[col] = df_category[col].apply(lambda x: f"{float(x):.2f}%")

                dataframes[category] = df_category

        # Save to Excel
        output_file = 'AWS_Service_Category_Analysis.xlsx'
        print(f"\nSaving analysis to {output_file}...")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for category, df in dataframes.items():
                # Clean sheet name
                sheet_name = f'{category} Analysis'.replace('/', '_').replace('\\', '_')[:31]
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Format worksheet
                worksheet = writer.sheets[sheet_name]
                
                # Format headers
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', wrap_text=True)

                # Adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

        print(f"\nAnalysis complete! File saved to {output_file}")
        print("\nAnalysis includes sheets for:")
        for category in service_categories.keys():
            print(f"- {category}")
        
        return dataframes

    except Exception as e:
        print(f"\nError in analysis: {str(e)}")
        print("\nStack trace:")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    dfs = analyze_aws_services()
