
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Color
import os
import warnings
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Color
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, Rule
from openpyxl.styles.fills import PatternFill

# Suppress specific warnings
warnings.filterwarnings('ignore', category=FutureWarning)

def get_service_domains():
    """Return list of service domains and their associated services from the input file"""
    return {
        'Analytics Services': [
            'EMR', 'Athena', 'Analytics', 'Glue', 'QuickSight', 'Redshift',
            'Kinesis Analytics', 'Lake Formation', 'Data Pipeline', 'OpenSearch',
            'Open Source Data Analytics', 'Glue & Lake Formation'
        ],
        'Automation and Messaging Group': [
            'SNS', 'SQS', 'EventBridge', 'MQ', 'Step Functions', 'CW Events',
            'Simple Queue Service', 'Simple Notification Service', 'Amazon MQ',
            'AWS Step Functions', 'Pinwheel'
        ],
        'Compute Services': [
            'EC2', 'EC2 - Linux', 'EC2 - Windows', 'EC2 - Bandwidth',
            'Elastic Load Balancing', 'Load Balancer', 'NAT Gateway',
            'Public IPv4 Address', 'Containers', 'EKS', 'ECR',
            'Elastic Load Balancing - ALB', 'Elastic Load Balancing - NLB',
            'Elastic Load Balancing - CLB'
        ],
        'DB Services': [
            'RDS', 'DynamoDB', 'Elasticache', 'Aurora', 'DDB',
            'Relational Database Services', 'Aurora MySQL', 'Aurora PostgreSQL',
            'RDS MySQL', 'RDS PostgreSQL', 'RDS MariaDB', 'RDS Storage',
            'Aurora MySQL Serverless v2'
        ],
        'Developer Tools': [
            'CodeBuild', 'CodePipeline', 'CodeDeploy', 'CodeCommit', 'Cloud9',
            'CodeStar', 'CodeArtifact', 'CodeGuru', 'X-Ray', 'CloudFormation',
            'AWS CodeSuite and IDEs', 'Infra-as-Code'
        ],
        'Edge': [
            'CloudFront', 'Route53', 'PerimeterProtection', 'Cloud Front',
            'Global Accelerator', 'Route 53'
        ],
        'Identity Services': [
            'IAM', 'Directory Service', 'Cognito', 'SSO', 'Single Sign-On',
            'Identity Center', 'Resource Access Manager', 'AWS Organizations',
            'AWS Directory Service'
        ],
        'Machine Learning & Deep Learning': [
            'SageMaker', 'Rekognition', 'Comprehend', 'Textract', 'Polly',
            'Transcribe', 'Bedrock', 'Lex', 'Translate', 'QuickSight',
            'Total Marketing Intelligence', 'Amazon Q Business'
        ],
        'Monitoring Services': [
            'CloudWatch', 'CW', 'CloudTrail', 'Config', 'Managed Services',
            'AWS Health', 'Systems Manager', 'SSM', 'Control Tower',
            'AWS Managed Services', 'Service Catalog', 'License Manager',
            'AWS Systems Manager', 'AWS CloudWatch', 'AWS Config'
        ],
        'Management Tools': [
            'CloudTrail', 'Config', 'SSM', 'Systems Manager', 'CloudFormation',
            'Auto Scaling'
        ],
        'Security Services': [
            'GuardDuty', 'KMS', 'Inspector', 'Certificate', 'Firewall', 
            'WAF', 'Shield', 'Secrets Manager', 'ESS', 'Macie',
            'Security Hub', 'Detective', 'Network Firewall', 'ACM',
            'GuardDuty Services', 'ESS Overbridge', 'Key Management Service',
            'AWS Secrets Manager', 'Inspector Services Group'
        ],
        'Storage': [
            'S3', 'EBS', 'EFS', 'Storage', 'Glacier', 'Elastic Block Store',
            'FSx', 'Storage Gateway', 'Simple Storage Service', 'EFS Total',
            'AWS Backup', 'AWS Transfer'
        ],
        'Streaming Services': [
            'Kinesis', 'SNS', 'SQS', 'EventBridge', 'MQ', 'Streaming',
            'Message Queue', 'Data Streams', 'Firehose', 'Video Streams',
            'Amazon MSK', 'Kinesis Data Firehose'
        ]
    }


def create_overall_summary(cleaned_data, account_domain_spend):
    """Create overall summary dataframe"""
    total_accounts = len(cleaned_data)
    total_spend = sum(data['total'] for data in account_domain_spend.values())
    avg_yearly = total_spend / total_accounts if total_accounts > 0 else 0
    avg_monthly = avg_yearly / 12
    total_monthly = total_spend / 12
    yearly_spends = [data['total'] for data in account_domain_spend.values()]
    lowest_yearly = min(yearly_spends) if yearly_spends else 0
    highest_yearly = max(yearly_spends) if yearly_spends else 0
    
    summary_data = {
        'Metric': [
            'Total Number of Accounts',
            'Total AWS Spend (All Accounts, 12 months)',
            'Average Yearly Spend per Account',
            'Average Monthly Spend per Account',
            'Total Monthly Spend (All Accounts)',
            'Lowest Yearly Spend',
            'Highest Yearly Spend'
        ],
        'Value': [
            total_accounts,
            f"${total_spend:,.2f}",
            f"${avg_yearly:,.2f}",
            f"${avg_monthly:,.2f}",
            f"${total_monthly:,.2f}",
            f"${lowest_yearly:,.2f}",
            f"${highest_yearly:,.2f}"
        ]
    }
    
    return pd.DataFrame(summary_data)

def create_compute_services_summary(account_domain_spend):
    """Create Compute services summary dataframe"""
    compute_data = []
    compute_usage_pcts = []
    
    for account_name, data in account_domain_spend.items():
        total_spend = data['total']
        compute_spend = data['domains'].get('Compute Services', 0)
        compute_pct = (compute_spend / total_spend * 100) if total_spend > 0 else 0
        compute_usage_pcts.append((account_name, compute_pct))
        
        # Get the primary services and their spends
        primary_services = get_primary_compute_services(cleaned_data.get(account_name))
        
        # Calculate core compute percentage (EC2 Linux + EC2 Windows)
        core_compute_spend = 0
        if cleaned_data.get(account_name) is not None:
            for _, row in cleaned_data[account_name].iterrows():
                service = str(row.iloc[0])
                if 'EC2 - Linux' in service or 'EC2 - Windows' in service:
                    core_compute_spend += clean_amount(row.iloc[1])
        
        core_compute_pct = (core_compute_spend / total_spend * 100) if total_spend > 0 else 0
        
        compute_data.append({
            'Customer Name': account_name,
            'Total Spend': f"${total_spend:,.2f}",
            'Compute Services Total': f"${compute_spend:,.2f}",
            'Compute Services %': f"{compute_pct:.2f}%",
            'Core Compute %': f"{core_compute_pct:.2f}%",
            'Primary Compute Services': primary_services
        })
    
    highest_compute = max(compute_usage_pcts, key=lambda x: x[1]) if compute_usage_pcts else ('None', 0)
    avg_compute = sum(pct for _, pct in compute_usage_pcts) / len(compute_usage_pcts) if compute_usage_pcts else 0
    
    # Create summary section
    compute_summary_rows = pd.DataFrame({
        'Customer Name': [
            'Compute Services Analysis Summary',
            '',
            'Total Customers Using Compute',
            'Highest Compute Usage',
            'Average Compute Usage',
            'Most Common Services',
            ''
        ],
        'Total Spend': [
            '',
            '',
            str(sum(1 for _, pct in compute_usage_pcts if pct > 0)),
            f"{highest_compute[1]:.2f}% ({highest_compute[0]})",
            f"{avg_compute:.2f}%",
            get_most_common_compute_services(cleaned_data),
            ''
        ],
        'Compute Services Total': [''] * 7,
        'Compute Services %': [''] * 7,
        'Core Compute %': [''] * 7,
        'Primary Compute Services': [''] * 7
    })
    
    # Combine summary and details
    df_compute = pd.concat([compute_summary_rows, pd.DataFrame(compute_data)], ignore_index=True)
    
    return df_compute

def create_storage_services_summary(account_domain_spend):
    """Create Storage services summary dataframe"""
    storage_data = []
    storage_usage_pcts = []
    
    for account_name, data in account_domain_spend.items():
        total_spend = data['total']
        storage_spend = data['domains'].get('Storage', 0)
        storage_pct = (storage_spend / total_spend * 100) if total_spend > 0 else 0
        storage_usage_pcts.append((account_name, storage_pct))
        
        storage_data.append({
            'Customer Name': account_name,
            'Total Spend': f"${total_spend:,.2f}",
            'Storage Services Total': f"${storage_spend:,.2f}",
            'Storage Services %': f"{storage_pct:.2f}%",
            'Primary Storage Services': get_primary_storage_services(cleaned_data.get(account_name))
        })
    
    highest_storage = max(storage_usage_pcts, key=lambda x: x[1]) if storage_usage_pcts else ('None', 0)
    avg_storage = sum(pct for _, pct in storage_usage_pcts) / len(storage_usage_pcts) if storage_usage_pcts else 0
    
    storage_summary_rows = pd.DataFrame({
        'Customer Name': [
            'Storage Services Analysis Summary',
            '',
            'Total Customers Using Storage',
            'Highest Storage Usage',
            'Average Storage Usage',
            'Most Common Services',
            ''
        ],
        'Total Spend': [
            '',
            '',
            str(sum(1 for _, pct in storage_usage_pcts if pct > 0)),
            f"{highest_storage[1]:.2f}% ({highest_storage[0]})",
            f"{avg_storage:.2f}%",
            get_most_common_storage_services(cleaned_data),
            ''
        ],
        'Storage Services Total': [''] * 7,
        'Storage Services %': [''] * 7,
        'Primary Storage Services': [''] * 7
    })
    
    return pd.concat([storage_summary_rows, pd.DataFrame(storage_data)], ignore_index=True)

def get_database_service_types():
    """Define the categorization of database services"""
    return {
        'managed': [
            'RDS MySQL',
            'RDS PostgreSQL',
            'RDS MariaDB',
            'RDS Oracle',
            'RDS Storage',
            'RDS SQL Server',
            'RDS Instances',
            'Relational Database Services (RDS) - Aurora MySQL',  # Regular Aurora MySQL is managed
            'Relational Database Services (RDS) - Aurora PostgreSQL'  # Regular Aurora PostgreSQL is managed
        ],
        'serverless': [
            'Aurora MySQL Serverless',
            'Aurora PostgreSQL Serverless',
            'Aurora Serverless v2',
            'DynamoDB',
            'DDB',
            'Elasticache',
            'ElastiCache',
            'DocumentDB'
        ]
    }


def get_primary_database_services(df):
    """Identify primary database services used by account and categorize them"""
    if df is None:
        return "No data"
    
    # Get service categorization
    db_types = get_database_service_types()
    managed_db_services = db_types['managed']
    serverless_db_services = db_types['serverless']
    
    managed_services = []
    serverless_services = []
    
    for _, row in df.iterrows():
        service = str(row.iloc[0])
        spend = clean_amount(row.iloc[1])
        
        if spend > 0:
            # Check for serverless services first (more specific matches)
            if any(s.lower() in service.lower() for s in serverless_db_services):
                serverless_services.append((service, spend))
            # Then check for managed services
            elif any(s.lower() in service.lower() for s in managed_db_services):
                managed_services.append((service, spend))
    
    if not managed_services and not serverless_services:
        return "None"
    
    # Sort services by spend within each category
    managed_services.sort(key=lambda x: x[1], reverse=True)
    serverless_services.sort(key=lambda x: x[1], reverse=True)
    
    # Format the output
    result = []
    
    if managed_services:
        managed_str = "Managed: " + "; ".join(s[0] for s in managed_services[:2])
        result.append(managed_str)
        
    if serverless_services:
        serverless_str = "Serverless: " + "; ".join(s[0] for s in serverless_services[:2])
        result.append(serverless_str)
    
    return " | ".join(result)

def create_database_services_summary(account_domain_spend):
    """Create Database services summary dataframe"""
    db_data = []
    db_usage_pcts = []
    
    # Track serverless and managed usage
    serverless_users = set()
    managed_users = set()
    total_accounts = 0
    
    for account_name, data in account_domain_spend.items():
        total_accounts += 1
        total_spend = data['total']
        db_spend = data['domains'].get('DB Services', 0)
        db_pct = (db_spend / total_spend * 100) if total_spend > 0 else 0
        db_usage_pcts.append((account_name, db_pct))
        
        # Calculate service type totals
        if cleaned_data.get(account_name) is not None:
            has_managed = False
            has_serverless = False
            
            for _, row in cleaned_data[account_name].iterrows():
                service = str(row.iloc[0])
                spend = clean_amount(row.iloc[1])
                
                if spend > 0:
                    # Explicitly check for serverless services
                    if (‘aurora' in service.lower() and 'serverless' in service.lower()) or \
                       'dynamodb' in service.lower() or \
                       'ddb' in service.lower() or \
                       'elasticache' in service.lower() or \
                       'documentdb' in service.lower():
                        has_serverless = True
                        print(f"Serverless service found for {account_name}: {service}")  # Debug log
                    
                    # Check for managed services
                    elif 'rds' in service.lower() or \
                         ('aurora' in service.lower() and 'serverless' not in service.lower()):
                        has_managed = True
                        print(f"Managed service found for {account_name}: {service}")  # Debug log
            
            if has_serverless:
                serverless_users.add(account_name)
            if has_managed:
                managed_users.add(account_name)
        
        primary_services = get_primary_database_services(cleaned_data.get(account_name))
        
        db_data.append({
            'Customer Name': account_name,
            'Total Spend': f"${total_spend:,.2f}",
            'Database Services Total': f"${db_spend:,.2f}",
            'Database Services %': f"{db_pct:.2f}%",
            'Primary Database Services': primary_services
        })
    
    highest_db = max(db_usage_pcts, key=lambda x: x[1]) if db_usage_pcts else ('None', 0)
    avg_db = sum(pct for _, pct in db_usage_pcts) / len(db_usage_pcts) if db_usage_pcts else 0
    
    # Print debug information
    print("\nServerless Users:", sorted(serverless_users))
    print("Managed Users:", sorted(managed_users))
    
    # Calculate percentages
    serverless_pct = (len(serverless_users) / total_accounts * 100) if total_accounts > 0 else 0
    managed_pct = (len(managed_users) / total_accounts * 100) if total_accounts > 0 else 0
    
    # Format serverless users list
    serverless_accounts_str = ', '.join(sorted(serverless_users)) if serverless_users else 'None'
    managed_accounts_str = ', '.join(sorted(managed_users)) if managed_users else 'None'
    
    serverless_users_str = f"{serverless_pct:.1f}% ({len(serverless_users)} customers): {serverless_accounts_str}"
    managed_users_str = f"{managed_pct:.1f}% ({len(managed_users)} customers): {managed_accounts_str}"
    
    # Create summary section
    db_summary_rows = pd.DataFrame({
        'Customer Name': [
            'Database Services Analysis Summary',
            '',
            'Serverless Database Usage',
            'Managed Database Usage',
            '',
            'Total Customers Using Databases',
            'Highest Database Usage',
            'Average Database Usage',
            'Most Common Services',
            ''
        ],
        'Total Spend': [
            '',
            '',
            serverless_users_str,
            managed_users_str,
            '',
            str(sum(1 for _, pct in db_usage_pcts if pct > 0)),
            f"{highest_db[1]:.2f}% ({highest_db[0]})",
            f"{avg_db:.2f}%",
            get_most_common_database_services(cleaned_data),
            ''
        ],
        'Database Services Total': [''] * 10,
        'Database Services %': [''] * 10,
        'Primary Database Services': [''] * 10
    })
    
    # Sort the detail rows by Database Services % (descending)
    df_details = pd.DataFrame(db_data)
    df_details['Sort Value'] = df_details['Database Services %'].apply(
        lambda x: float(x.replace('%', '')) if x != '' else 0
    )
    df_details = df_details.sort_values('Sort Value', ascending=False)
    df_details = df_details.drop('Sort Value', axis=1)
    
    # Combine summary and sorted details
    df_database = pd.concat([db_summary_rows, df_details], ignore_index=True)
    
    return df_database



# Helper functions to identify primary services
def get_primary_compute_services(df):
    """Identify primary compute services used by account"""
    if df is None:
        return "No data"
    
    compute_services = [
        'EC2', 'ECS', 'EKS', 'Lambda', 'Elastic Load Balancing',
        'Auto Scaling', 'Elastic Beanstalk'
    ]
    
    services = []
    for _, row in df.iterrows():
        service = str(row.iloc[0])
        if any(cs.lower() in service.lower() for cs in compute_services):
            spend = clean_amount(row.iloc[1])
            if spend > 0:
                services.append((service, spend))
    
    if not services:
        return "None"
    
    # Return top 3 by spend
    top_services = sorted(services, key=lambda x: x[1], reverse=True)[:3]
    return "; ".join(f"{s[0]}" for s in top_services)

def get_primary_storage_services(df):
    """Identify primary storage services used by account"""
    if df is None:
        return "No data"
    
    storage_services = [
        'S3', 'EBS', 'EFS', 'FSx', 'Storage Gateway',
        'Glacier', 'Backup'
    ]
    
    services = []
    for _, row in df.iterrows():
        service = str(row.iloc[0])
        if any(ss.lower() in service.lower() for ss in storage_services):
            spend = clean_amount(row.iloc[1])
            if spend > 0:
                services.append((service, spend))
    
    if not services:
        return "None"
    
    top_services = sorted(services, key=lambda x: x[1], reverse=True)[:3]
    return "; ".join(f"{s[0]}" for s in top_services)

def get_primary_database_services(df):
    """Identify primary database services used by account"""
    if df is None:
        return "No data"
    
    db_services = [
        'RDS', 'Aurora', 'DynamoDB', 'ElastiCache',
        'Redshift', 'DocumentDB', 'Neptune'
    ]
    
    services = []
    for _, row in df.iterrows():
        service = str(row.iloc[0])
        if any(ds.lower() in service.lower() for ds in db_services):
            spend = clean_amount(row.iloc[1])
            if spend > 0:
                services.append((service, spend))
    
    if not services:
        return "None"
    
    top_services = sorted(services, key=lambda x: x[1], reverse=True)[:3]
    return "; ".join(f"{s[0]}" for s in top_services)

# Helper functions to get most common services across all accounts
def get_most_common_compute_services(cleaned_data):
    """Get most commonly used compute services across all accounts"""
    service_counts = {}
    for df in cleaned_data.values():
        services = get_primary_compute_services(df).split("; ")
        for service in services:
            if service not in ["None", "No data"]:
                service_counts[service] = service_counts.get(service, 0) + 1
    
    if not service_counts:
        return "None"
    
    top_services = sorted(service_counts.items(), key=lambda x: x[1], reverse=True)[:3]
    return "; ".join(f"{s[0]}" for s in top_services)

def get_most_common_storage_services(cleaned_data):
    """Get most commonly used storage services across all accounts"""
    service_counts = {}
    for df in cleaned_data.values():
        services = get_primary_storage_services(df).split("; ")
        for service in services:
            if service not in ["None", "No data"]:
                service_counts[service] = service_counts.get(service, 0) + 1
    
    if not service_counts:
        return "None"
    
    top_services = sorted(service_counts.items(), key=lambda x: x[1], reverse=True)[:3]
    return "; ".join(f"{s[0]}" for s in top_services)

def get_most_common_database_services(cleaned_data):
    """Get most commonly used database services across all accounts"""
    service_counts = {}
    for df in cleaned_data.values():
        if df is not None:
            for _, row in df.iterrows():
                service = str(row.iloc[0])
                spend = clean_amount(row.iloc[1])
                if spend > 0:
                    if any(db_service in service for db_service in ['RDS', 'Aurora', 'DynamoDB', 'ElastiCache', 'DocumentDB']):
                        service_counts[service] = service_counts.get(service, 0) + 1
    
    if not service_counts:
        return "None"
    
    # Get top 3 most common services
    top_services = sorted(service_counts.items(), key=lambda x: x[1], reverse=True)[:3]
    return "; ".join(service[0] for service in top_services)

def create_ml_services_summary(account_domain_spend):
    """Create ML services summary dataframe"""
    # Create ML services detail rows
    ml_data = []
    ml_usage_pcts = []
    
    for account_name, data in account_domain_spend.items():
        total_spend = data['total']
        ml_spend = data['domains'].get('Machine Learning & Deep Learning', 0)
        ml_pct = (ml_spend / total_spend * 100) if total_spend > 0 else 0
        ml_usage_pcts.append((account_name, ml_pct))
        
        ml_data.append({
            'Customer Name': account_name,
            'Total Spend': f"${total_spend:,.2f}",
            'ML Services Total': f"${ml_spend:,.2f}",
            'ML Services %': f"{ml_pct:.2f}%"
        })
    
    highest_ml = max(ml_usage_pcts, key=lambda x: x[1]) if ml_usage_pcts else ('None', 0)
    avg_ml = sum(pct for _, pct in ml_usage_pcts) / len(ml_usage_pcts) if ml_usage_pcts else 0
    
    # Create summary section
    ml_summary_rows = pd.DataFrame({
        'Customer Name': [
            'ML Services Analysis Summary',
            '',
            'Total Customers Using ML Services',
            'Highest ML Usage',
            'Average ML Usage',
            ''
        ],
        'Total Spend': [
            '',
            '',
            str(sum(1 for _, pct in ml_usage_pcts if pct > 0)),
            f"{highest_ml[1]:.2f}% ({highest_ml[0]})",
            f"{avg_ml:.2f}%",
            ''
        ],
        'ML Services Total': [''] * 6,
        'ML Services %': [''] * 6
    })
    
    # Combine summary and details
    df_ml_details = pd.DataFrame(ml_data)
    df_ml_combined = pd.concat([ml_summary_rows, df_ml_details], ignore_index=True)
    
    return df_ml_combined

def clean_excel_data(excel_file_path):
    """Clean and standardize Excel data"""
    excel_file = pd.ExcelFile(excel_file_path)
    cleaned_sheets = {}
    
    for sheet_name in excel_file.sheet_names:
        try:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
            if df.empty or df.dropna().empty:
                print(f"Sheet '{sheet_name}' is empty - skipping")
                continue
            
            print(f"\nProcessing sheet: {sheet_name}")
            print(f"Original shape: {df.shape}")
            
            new_columns = []
            for col in df.columns:
                if isinstance(col, str):
                    if 'Service group' in col:
                        new_columns.append('Service group')
                    elif '12 months total' in col:
                        new_columns.append('Last 12 months total')
                    else:
                        new_columns.append(col)
                elif isinstance(col, datetime):
                    new_columns.append(col.strftime('%b-%y'))
                else:
                    new_columns.append(col)
            
            df.columns = new_columns
            df = df.replace('-', 0)
            
            if not df.empty:
                cleaned_sheets[sheet_name] = df
                print(f"Shape after cleaning: {df.shape}")
                print("\nColumn names after cleaning:")
                print(df.columns.tolist())
            
        except Exception as e:
            print(f"Error processing sheet '{sheet_name}': {str(e)}")
            continue
    
    return cleaned_sheets

def clean_amount(value):
    """Clean and convert monetary values to float"""
    try:
        if pd.isna(value):
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            if '(free tier)' in value:
                return 0.0
            cleaned = value.replace('US$', '').replace('$', '').replace(',', '').replace(' ', '')
            if cleaned == '-' or cleaned == '':
                return 0.0
            return float(cleaned)
        return 0.0
    except Exception as e:
        print(f"Warning: Could not convert '{value}' to float. Using 0.0 instead.")
        return 0.0

def get_service_domain(service_name, domain_mapping):
    """Match a service name to its domain"""
    service_name = str(service_name).lower()
    for domain, services in domain_mapping.items():
        if any(service.lower() in service_name for service in services):
            return domain
    return None

def get_account_total_spend(df):
    """Get total spend for an account from the first data row"""
    try:
        # Skip empty rows and headers
        for idx, row in df.iterrows():
            # Find the first non-empty row after headers
            if pd.notna(row.iloc[0]) and isinstance(row.iloc[0], str):
                if not row.iloc[0].startswith('Service group'):
                    # Get value from "Last 12 months total" column
                    total_spend_str = row.iloc[1]  # Column index 1 contains "Last 12 months total"
                    # Clean and convert the value
                    if isinstance(total_spend_str, str):
                        total_spend = float(total_spend_str.replace('US$', '').replace(',', ''))
                    else:
                        total_spend = float(total_spend_str)
                    account_name = row.iloc[0]
                    print(f"Found total spend for {account_name}: ${total_spend:,.2f}")
                    return total_spend
        return 0.0
    except Exception as e:
        print(f"Error getting account total spend: {str(e)}")
        return 0.0

def analyze_service_domains(cleaned_data):
    """Analyze spending by service domain"""
    domain_mapping = get_service_domains()
    domain_totals = {domain: 0 for domain in domain_mapping.keys()}
    domain_accounts = {domain: set() for domain in domain_mapping.keys()}
    account_domain_spend = {}
    account_totals = {}

    # First pass - get account totals 
    for sheet_name, df in cleaned_data.items():
        account_totals[sheet_name] = get_account_total_spend(df)
        account_domain_spend[sheet_name] = {
            'total': account_totals[sheet_name],
            'domains': {domain: 0 for domain in domain_mapping.keys()}
        }

    # Second pass - process service domains and their spends
    for sheet_name, df in cleaned_data.items():
        try:
            # Track when we're in a domain section
            current_domain = None
            
            for idx, row in df.iterrows():
                service_name = str(row.iloc[0])
                
                # Skip empty rows and "View AWS service" rows
                if pd.isna(service_name) or 'View AWS service' in service_name:
                    continue

                # Check if this row is a domain header
                for domain, services in domain_mapping.items():
                    # Check if the service name exactly matches a domain name
                    if service_name == domain:
                        spend = clean_amount(row.iloc[1])  # Get spend from column 2
                        if spend > 0:
                            domain_totals[domain] += spend
                            domain_accounts[domain].add(sheet_name)
                            account_domain_spend[sheet_name]['domains'][domain] = spend
                        break

        except Exception as e:
            print(f"Error processing {sheet_name}: {str(e)}")
            continue

    # Create summary
    summary_data = []
    total_spend = sum(account_totals.values())

    for domain in domain_mapping.keys():
        spend = domain_totals[domain]
        unique_accounts = len(domain_accounts[domain])
        avg_spend = spend / unique_accounts if unique_accounts > 0 else 0
        percentage = (spend / total_spend * 100) if total_spend > 0 else 0

        summary_data.append({
            'Service Domain': domain,
            'Total Spend': f"${spend:,.2f}",
            'Number of Accounts': unique_accounts,
            'Average Spend': f"${avg_spend:,.2f}", 
            'Percentage of Total': f"{percentage:.2f}%",
            'Active Accounts': ', '.join(sorted(domain_accounts[domain])) if unique_accounts > 0 else 'None'
        })

    df_summary = pd.DataFrame(summary_data)
    return df_summary.sort_values(
        'Total Spend',
        key=lambda x: pd.to_numeric(x.str.replace('$', '').replace(',', ''), errors='coerce'),
        ascending=False
    ), account_domain_spend



def create_comparative_analysis(cleaned_data, account_domain_spend):
    """Create comparative analysis between accounts"""
    domain_mapping = get_service_domains()
    comparative_data = []
    monthly_trends = {}
    
    for sheet_name, df in cleaned_data.items():
        try:
            total_spend = account_domain_spend[sheet_name]['total']  # Use stored total
            monthly_spends = []
            
            # Calculate monthly spends
            for month in range(2, 14):
                monthly_total = 0
                for idx, row in df.iterrows():
                    service_name = str(row.iloc[0])
                    if get_service_domain(service_name, domain_mapping):
                        monthly_total += clean_amount(row.iloc[month])
                monthly_spends.append(monthly_total)
            
            # Get domain spends
            domain_spends = account_domain_spend[sheet_name]['domains']
            if domain_spends:
                max_domain = max(domain_spends.items(), key=lambda x: x[1])
                domain_concentration = (max_domain[1] / total_spend * 100) if total_spend > 0 else 0
            else:
                max_domain = ('None', 0)
                domain_concentration = 0
            
            # Calculate month-over-month changes
            mom_changes = []
            for i in range(1, len(monthly_spends)):
                if monthly_spends[i-1] != 0:
                    change = ((monthly_spends[i] - monthly_spends[i-1]) / monthly_spends[i-1]) * 100
                else:
                    change = 0
                mom_changes.append(change)
            
            comparative_data.append({
                'Account': sheet_name,
                'Total Annual Spend': total_spend,
                'Avg Monthly Spend': total_spend / 12 if total_spend > 0 else 0,
                'Highest Spend Domain': max_domain[0],
                'Domain Spend': max_domain[1],
                'Domain Concentration %': domain_concentration,
                'Max Monthly Spend': max(monthly_spends) if monthly_spends else 0,
                'Min Monthly Spend': min(monthly_spends) if monthly_spends else 0,
                'Spend Volatility': np.std(monthly_spends) if monthly_spends else 0,
                'Avg MoM Change %': np.mean(mom_changes) if mom_changes else 0
            })
            
            monthly_trends[sheet_name] = monthly_spends
            
        except Exception as e:
            print(f"Error processing {sheet_name} in comparative analysis: {str(e)}")
            continue
    
    df_comparative = pd.DataFrame(comparative_data)
    
    # Format columns
    for col in df_comparative.columns:
        if col.endswith('Spend'):
            df_comparative[col] = df_comparative[col].apply(lambda x: f"${x:,.2f}")
        elif col.endswith('%'):
            df_comparative[col] = df_comparative[col].apply(lambda x: f"{x:.2f}%")
    
    # Create monthly trends DataFrame
    month_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    df_trends = pd.DataFrame(monthly_trends).T
    df_trends.columns = month_labels
    
    for col in df_trends.columns:
        df_trends[col] = df_trends[col].apply(lambda x: f"${x:,.2f}")
    
    return df_comparative, df_trends
   

def create_cumulative_report(cleaned_data, account_domain_spend):
    """Create detailed cumulative analysis report"""
    domain_mapping = get_service_domains()
    domains = list(domain_mapping.keys())
    cumulative_data = []
    
    # Calculate overall totals using stored account totals
    total_spend = sum(data['total'] for data in account_domain_spend.values())
    num_accounts = len(account_domain_spend)
    
    # Create summary row
    summary_row = {
        'Account': 'TOTAL ALL ACCOUNTS',
        'Total Annual Spend': total_spend,
        'Average Monthly Spend': total_spend / 12 if total_spend > 0 else 0,
        'Number of Active Domains': 0,
        'Most Used Domain': '',
        'Domain Distribution Score': 0
    }
    
    # Calculate domain totals
    domain_totals = {domain: 0 for domain in domains}
    for account_data in account_domain_spend.values():
        for domain, spend in account_data['domains'].items():
            domain_totals[domain] += spend
    
    # Add domain-specific fields to summary
    for domain in domains:
        summary_row[f'{domain} Total'] = domain_totals[domain]
        summary_row[f'{domain} Avg'] = domain_totals[domain] / num_accounts if num_accounts > 0 else 0
    
    # Find most used domain for summary
    if domain_totals:
        most_used = max(domain_totals.items(), key=lambda x: x[1])
        summary_row['Most Used Domain'] = f"{most_used[0]} (${most_used[1]:,.2f})"
        summary_row['Number of Active Domains'] = sum(1 for v in domain_totals.values() if v > 0)
    
    cumulative_data.append(summary_row)
    
    # Process each account
    for account, data in account_domain_spend.items():
        account_total = data['total']
        domain_spends = data['domains']
        active_domains = sum(1 for v in domain_spends.values() if v > 0)
        
        # Calculate domain distribution score
        if account_total > 0:
            proportions = [spend/account_total for spend in domain_spends.values() if spend > 0]
            distribution_score = 1 - np.std(proportions) if proportions else 0
        else:
            distribution_score = 0
        
        account_row = {
            'Account': account,
            'Total Annual Spend': account_total,
            'Average Monthly Spend': account_total / 12,
            'Number of Active Domains': active_domains,
            'Most Used Domain': '',
            'Domain Distribution Score': distribution_score * 100
        }
        
        # Find account's most used domain
        if domain_spends:
            most_used = max(domain_spends.items(), key=lambda x: x[1])
            account_row['Most Used Domain'] = f"{most_used[0]} (${most_used[1]:,.2f})"
        
        # Add domain-specific spending and comparisons
        for domain in domains:
            spend = domain_spends.get(domain, 0)
            avg_spend = summary_row[f'{domain} Avg']
            account_row[f'{domain} Spend'] = spend
            account_row[f'{domain} vs Avg'] = ((spend - avg_spend) / avg_spend * 100) if avg_spend > 0 else 0
        
        cumulative_data.append(account_row)
    
    df_cumulative = pd.DataFrame(cumulative_data)
    
    # Format the DataFrame
    currency_cols = ['Total Annual Spend', 'Average Monthly Spend'] + \
                   [f'{domain} Spend' for domain in domains]
    percentage_cols = [f'{domain} vs Avg' for domain in domains] + ['Domain Distribution Score']
    
    for col in currency_cols:
        if col in df_cumulative.columns:
            df_cumulative[col] = df_cumulative[col].apply(lambda x: f"${x:,.2f}")
    
    for col in percentage_cols:
        if col in df_cumulative.columns:
            df_cumulative[col] = df_cumulative[col].apply(lambda x: f"{x:.1f}%" if isinstance(x, (int, float)) else x)
    
    return df_cumulative

def calculate_category_distribution_score(row, categories):
    """Calculate the distribution score for service categories"""
    percentages = []
    for category in categories:
        try:
            if f"{category} %" in row:
                percentage = float(str(row[f"{category} %"]).replace('%', ''))
                if percentage > 0:
                    percentages.append(percentage)
        except:
            continue
    
    if percentages:
        return round(100 - np.std(percentages), 2)
    return 0.0


def clean_amount_string(val):
    """Helper function to clean and convert amount strings to float"""
    try:
        if pd.isna(val):
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        val_str = str(val).lower()
        if '(free tier)' in val_str:
            return 0.0
        # Remove currency symbols, commas and spaces
        cleaned = val_str.replace('us$', '').replace('$', '').replace(',', '').replace(' ', '')
        if cleaned == '-' or cleaned == '':
            return 0.0
        return float(cleaned)
    except:
        print(f"Warning: Could not convert '{val}' to float. Using 0.0")
        return 0.0

def analyze_monthly_patterns(cleaned_data, account_domain_spend):
    """Analyze monthly spending patterns using direct data from MyPodData.xlsx"""
    monthly_patterns = []
    
    for sheet_name, df in cleaned_data.items():
        print(f"\nAnalyzing patterns for {sheet_name}...")
        account_patterns = {
            'Account': sheet_name,
            'Service Patterns': {}
        }
        
        # Skip the header rows and get service rows
        service_rows = df[df['Service group'].notna()]
        
        # Get monthly columns (assuming they are in columns 3-14)
        monthly_cols = df.columns[2:14]  # First 2 columns are service group and total
        
        for index, row in service_rows.iterrows():
            service_name = row['Service group']
            if isinstance(service_name, str) and not 'View AWS service' in service_name:
                monthly_spend = []
                
                # Get monthly values directly from the Excel data
                for col in monthly_cols:
                    value = row[col]
                    if isinstance(value, str):
                        value = clean_amount(value)
                    monthly_spend.append(float(value) if value else 0.0)
                
                if sum(monthly_spend) > 0:
                    mean_spend = np.mean(monthly_spend)
                    std_spend = np.std(monthly_spend)
                    
                    # Calculate anomalies
                    anomalies = []
                    for month_idx, spend in enumerate(monthly_spend):
                        z_score = (spend - mean_spend) / std_spend if std_spend > 0 else 0
                        if abs(z_score) > 2:
                            anomalies.append({
                                'Month': month_idx + 1,
                                'Spend': spend,
                                'Z-score': z_score
                            })
                    
                    # Calculate pattern info
                    pattern_info = {
                        'Monthly Spend': monthly_spend,
                        'Average': mean_spend,
                        'Std Dev': std_spend,
                        'Anomalies': anomalies,
                        'Trend': calculate_trend_pattern(monthly_spend),
                        'Max Month': np.argmax(monthly_spend) + 1,
                        'Min Month': np.argmin(monthly_spend) + 1,
                        'Volatility': std_spend / mean_spend if mean_spend > 0 else 0
                    }
                    
                    account_patterns['Service Patterns'][service_name] = pattern_info
        
        monthly_patterns.append(account_patterns)
    
    return monthly_patterns


def calculate_trend_pattern(monthly_data):
    """Identify the trend pattern in monthly spending"""
    if not monthly_data or len(monthly_data) < 2:
        return {
            'type': 'insufficient_data',
            'description': 'Insufficient data',
            'trend_slope': 0,
            'coefficient_of_variation': 0,
            'half_year_change': 0
        }
    
    try:
        mean_spend = np.mean(monthly_data)
        std_spend = np.std(monthly_data)
        cv = std_spend / mean_spend if mean_spend > 0 else 0
        
        x = np.arange(len(monthly_data))
        slope, _ = np.polyfit(x, monthly_data, 1)
        
        first_half_avg = np.mean(monthly_data[:6])
        second_half_avg = np.mean(monthly_data[6:])
        half_year_change = ((second_half_avg - first_half_avg) / first_half_avg * 100) if first_half_avg > 0 else 0
        
        if cv > 0.5:
            pattern_type = 'highly_variable'
            description = f'Highly variable spending (CV: {cv:.2f})'
        elif cv > 0.25:
            pattern_type = 'moderately_variable'
            description = f'Moderately variable spending (CV: {cv:.2f})'
        elif slope > 0:
            if half_year_change > 50:
                pattern_type = 'sharp_increase'
                description = f'Sharp increasing trend (+{half_year_change:.1f}% in 6 months)'
            else:
                pattern_type = 'gradual_increase'
                description = f'Gradual increasing trend (+{half_year_change:.1f}% in 6 months)'
        elif slope < 0:
            if half_year_change < -50:
                pattern_type = 'sharp_decrease'
                description = f'Sharp decreasing trend ({half_year_change:.1f}% in 6 months)'
            else:
                pattern_type = 'gradual_decrease'
                description = f'Gradual decreasing trend ({half_year_change:.1f}% in 6 months)'
        else:
            pattern_type = 'stable'
            description = 'Stable spending pattern'
        
        return {
            'type': pattern_type,
            'description': description,
            'coefficient_of_variation': cv,
            'trend_slope': slope,
            'half_year_change': half_year_change
        }
    
    except Exception as e:
        print(f"Error in trend calculation: {str(e)}")
        return {
            'type': 'error',
            'description': 'Error in calculation',
            'trend_slope': 0,
            'coefficient_of_variation': 0,
            'half_year_change': 0
        }

def create_monthly_trends_report(monthly_patterns):
    """Create detailed monthly trends report"""
    month_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                   'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    
    trend_data = []
    
    if not monthly_patterns:
        print("No monthly patterns data available")
        return pd.DataFrame()
    
    for pattern in monthly_patterns:
        try:
            account = pattern['Account']
            
            for domain, info in pattern['Service Patterns'].items():
                monthly_spend = info['Monthly Spend']
                anomalies = info['Anomalies']
                trend = info['Trend']
                
                row_data = {
                    'Account': account,
                    'Service Domain': domain,
                    'Spending Pattern': trend['description'],
                    'Risk Level': calculate_risk_level(info),
                    'Volatility': f"{info['Volatility']*100:.1f}%",
                    'Notable Changes': format_anomalies(anomalies, month_labels),
                    'Average Monthly': f"${info['Average']:,.2f}",
                    'Trend Direction': 'Increasing' if trend.get('trend_slope', 0) > 0 else 'Decreasing'
                }
                
                for month_idx, spend in enumerate(monthly_spend):
                    row_data[month_labels[month_idx]] = f"${spend:,.2f}"
                
                trend_data.append(row_data)
                
        except Exception as e:
            print(f"Error processing trend data for {account}: {str(e)}")
            continue
    
    return pd.DataFrame(trend_data)

def format_anomalies(anomalies, month_labels):
    """Format anomaly information into readable text"""
    if not anomalies:
        return "No significant anomalies"
    
    anomaly_texts = []
    for anomaly in anomalies:
        try:
            month = month_labels[anomaly['Month']-1]
            direction = "spike" if anomaly['Z-score'] > 0 else "drop"
            magnitude = abs(anomaly['Z-score'])
            
            if magnitude > 3:
                severity = "extreme"
            elif magnitude > 2.5:
                severity = "significant"
            else:
                severity = "notable"
            
            amount = f"${anomaly['Spend']:,.2f}"
            anomaly_texts.append(f"{severity.title()} {direction} in {month} ({amount})")
        except Exception as e:
            print(f"Error formatting anomaly: {str(e)}")
            continue
    
    return "; ".join(anomaly_texts) if anomaly_texts else "Error processing anomalies"

def calculate_risk_level(pattern_info):
    """Calculate risk level based on spending pattern"""
    try:
        risk_score = 0
        
        # Factor 1: Volatility
        volatility = pattern_info.get('Volatility', 0)
        if volatility > 0.5:
            risk_score += 3
        elif volatility > 0.25:
            risk_score += 2
        elif volatility > 0.1:
            risk_score += 1
        
        # Factor 2: Number of anomalies
        num_anomalies = len(pattern_info.get('Anomalies', []))
        risk_score += min(num_anomalies, 3)
        
        # Factor 3: Trend pattern
        trend = pattern_info.get('Trend', {})
        if trend.get('type') in ['sharp_increase', 'sharp_decrease', 'highly_variable']:
            risk_score += 2
        elif trend.get('type') in ['gradual_increase', 'gradual_decrease', 'moderately_variable']:
            risk_score += 1
        
        # Determine risk level
        if risk_score >= 6:
            return 'High'
        elif risk_score >= 3:
            return 'Medium'
        return 'Low'
    
    except Exception as e:
        print(f"Error calculating risk level: {str(e)}")
        return 'Unknown'

def format_excel_with_highlights(worksheet, sheet_name=''):
    """Format Excel worksheet with conditional highlighting"""
    # Define colors
    header_color = "1F4E78"
    positive_var = "90EE90"
    negative_var = "FFB6C1"
    risk_high = "FF9999"
    risk_medium = "FFD699"
    summary_row = "E6E6E6"
    alt_row = "F5F5F5"
    
    # Define styles
    header_style = {
        'fill': PatternFill(start_color=header_color, end_color=header_color, fill_type="solid"),
        'font': Font(color="FFFFFF", bold=True),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }
    
    summary_style = {
        'fill': PatternFill(start_color=summary_row, end_color=summary_row, fill_type="solid"),
        'font': Font(bold=True),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }
    
    # Apply header formatting
    for cell in worksheet[1]:
        cell.fill = header_style['fill']
        cell.font = header_style['font']
        cell.alignment = header_style['alignment']
    
    # Format data rows
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), 2):
        for cell in row:
            # Default alignment
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Summary row formatting
            if row_idx == 2:
                cell.fill = summary_style['fill']
                cell.font = summary_style['font']
            
            # Alternate row colors
            elif row_idx % 2 == 0:
                cell.fill = PatternFill(start_color=alt_row, end_color=alt_row, fill_type="solid")
            
            # Risk level highlighting
            if 'Risk Level' in str(worksheet.cell(row=1, column=cell.column).value):
                if cell.value == 'High':
                    cell.fill = PatternFill(start_color=risk_high, end_color=risk_high, fill_type="solid")
                elif cell.value == 'Medium':
                    cell.fill = PatternFill(start_color=risk_medium, end_color=risk_medium, fill_type="solid")
            
            # Spending variations highlighting
            if isinstance(cell.value, str):
                # Percentage variations
                if 'vs Avg' in str(worksheet.cell(row=1, column=cell.column).value):
                    try:
                        value = float(cell.value.replace('%', '').replace('+', ''))
                        if value > 20:
                            cell.fill = PatternFill(start_color=positive_var, end_color=positive_var, fill_type="solid")
                        elif value < -20:
                            cell.fill = PatternFill(start_color=negative_var, end_color=negative_var, fill_type="solid")
                    except:
                        pass
                
                # Monthly spending variations
                elif cell.value.startswith('$'):
                    try:
                        value = float(cell.value.replace('$', '').replace(',', ''))
                        if value > 0:
                            row_values = [
                                float(str(c.value).replace('$', '').replace(',', ''))
                                for c in row
                                if isinstance(c.value, str) and c.value.startswith('$')
                            ]
                            if row_values:
                                row_avg = np.mean(row_values)
                                if value > row_avg * 1.5:
                                    cell.fill = PatternFill(start_color=positive_var, end_color=positive_var, fill_type="solid")
                                elif value < row_avg * 0.5:
                                    cell.fill = PatternFill(start_color=negative_var, end_color=negative_var, fill_type="solid")
                    except:
                        pass
    
    # Adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Add filters and freeze panes
    worksheet.freeze_panes = 'A2'
    worksheet.auto_filter.ref = worksheet.dimensions


def format_with_bold(value, average, is_percentage=False):
    """Format value with bold HTML if above average"""
    if isinstance(value, str):
        value = float(value.replace('$', '').replace(',', '').replace('%', ''))
    
    if is_percentage:
        formatted = f"{value:.2f}%"
    else:
        formatted = f"${value:,.2f}"
    
    if value > average:
        return f"<b>{formatted}</b>"
    return formatted

def get_top_categories(row, categories):
    """Get top 3 categories by percentage"""
    try:
        category_percentages = []
        for category in categories:
            pct_str = str(row[f"{category} %"])
            pct = float(pct_str.replace('<b>', '').replace('</b>', '').replace('%', ''))
            if pct > 0:
                category_percentages.append((category, pct))
        
        top_cats = sorted(category_percentages, key=lambda x: x[1], reverse=True)[:3]
        return '; '.join(f"{cat} ({pct:.1f}%)" for cat, pct in top_cats)
    except:
        return "None"

def calculate_industry_average(analysis_data):
    """Calculate industry average from individual account data"""
    avg_data = {
        'Customer Name': 'Industry Average',
        'Total Spend': f"${np.mean([float(d['Total Spend'].replace('$', '').replace(',', '')) for d in analysis_data]):,.2f}"
    }
    
    # Calculate averages for each category
    for key in analysis_data[0].keys():
        if key not in ['Customer Name', 'Total Spend', 'Top Categories', 'Category Distribution Score']:
            if 'Spend' in key:
                values = [float(d[key].replace('$', '').replace(',', '')) for d in analysis_data]
                avg_data[key] = f"${np.mean(values):,.2f}"
            elif '%' in key:
                values = [float(d[key].replace('%', '')) for d in analysis_data]
                avg_data[key] = f"{np.mean(values):.2f}%"
    
    # Calculate top categories for industry average
    spend_categories = [(cat.replace(' %', ''), float(avg_data[f"{cat.replace(' %', '')} %"].replace('%', '')))
                       for cat in avg_data.keys() if '% %' in key]
    top_cats = sorted(spend_categories, key=lambda x: x[1], reverse=True)[:3]
    avg_data['Top Categories'] = '; '.join(f"{cat} ({pct:.1f}%)" for cat, pct in top_cats if pct > 0)
    
    return avg_data


def calculate_category_distribution_score(row, categories):
    """Calculate the distribution score for service categories"""
    try:
        percentages = []
        for category in categories:
            if f"{category} %" in row:
                percentage = float(str(row[f"{category} %"]).replace('%', ''))
                if percentage > 0:
                    percentages.append(percentage)
        
        if percentages:
            # Higher score means more even distribution
            return round(100 - np.std(percentages), 2)
        return 0.0
    except Exception as e:
        print(f"Error calculating distribution score: {str(e)}")
        return 0.0


def calculate_distribution_score(row, percentage_cols):
    """Calculate a score representing how well-distributed the spending is across categories"""
    try:
        percentages = [float(str(row[col]).replace('%', '')) for col in percentage_cols]
        non_zero = [p for p in percentages if p > 0]
        if not non_zero:
            return "0.00"
        
        # Calculate Gini coefficient (0 = perfect distribution, 1 = complete concentration)
        gini = 1 - (1 / len(non_zero)) * (2 * sum((i+1) * v for i, v in enumerate(sorted(non_zero))) / sum(non_zero) - (len(non_zero) + 1))
        
        # Convert to distribution score (100 = perfect distribution, 0 = complete concentration)
        score = (1 - gini) * 100
        return f"{score:.2f}"
    except:
        return "0.00"

def get_top_categories(row, percentage_cols):
    """Get the top 3 categories by percentage"""
    try:
        categories = [(col.replace(' %', ''), float(str(row[col]).replace('%', ''))) 
                     for col in percentage_cols]
        top_cats = sorted(categories, key=lambda x: x[1], reverse=True)[:3]
        return '; '.join(f"{cat} ({pct:.1f}%)" for cat, pct in top_cats if pct > 0)
    except:
        return "None"


def format_sheet(worksheet, workbook, is_summary=False, is_ml_summary=False):
    """Format worksheet with specified styling"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
    
    # Define styles
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    white_font = Font(color='FFFFFF', bold=True)
    center_aligned = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_aligned = Alignment(horizontal='right', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Set column widths
    worksheet.column_dimensions['A'].width = 40  # Metric/Customer Name column
    for col in ['B', 'C', 'D']:
        worksheet.column_dimensions[col].width = 25  # Value/other columns

    if is_summary:
        # Format header row
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = center_aligned
            cell.border = border

        # Format value column for currency
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet[f'B{row}']
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('$'):
                cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                cell.alignment = right_aligned

    elif is_ml_summary:
        # Format the ML summary section headers
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = center_aligned
            cell.border = border

        # Format the details table header (row 9)
        for cell in worksheet[9]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = center_aligned
            cell.border = border

        # Format currency and percentage columns
        for row in range(2, worksheet.max_row + 1):
            # Format Total Spend column
            cell = worksheet[f'B{row}']
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('$'):
                cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                cell.alignment = right_aligned

            # Format ML Services Total column
            cell = worksheet[f'C{row}']
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('$'):
                cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                cell.alignment = right_aligned

            # Format ML Services % column
            cell = worksheet[f'D{row}']
            if cell.value and isinstance(cell.value, str) and cell.value.endswith('%'):
                cell.alignment = right_aligned


def create_service_group_summary(df_service_groups):
    """Create summary section for service group analysis"""
    def clean_percentage(value):
        """Clean percentage string and convert to float"""
        if isinstance(value, str):
            # Remove HTML tags and multiple percentages
            clean_val = value.replace('<b>', '').replace('</b>', '').split('%')[0]
            try:
                return float(clean_val)
            except ValueError:
                return 0.0
        return value

    def get_stats(column):
        """Get statistics for a column"""
        values = df_service_groups[column].apply(clean_percentage)
        avg = values.mean()
        max_val = values.max()
        min_val = values.min()
        max_account = df_service_groups.loc[values.idxmax(), 'Customer Name']
        min_account = df_service_groups.loc[values.idxmin(), 'Customer Name']
        return avg, max_val, min_val, max_account, min_account

    summary_data = []
    
    # Add header
    summary_data.append({
        'Metric': 'Service Group Analysis Summary',
        'Value': ''
    })
    summary_data.append({})  # Empty row
    
    # Process each service category
    categories = [
        'Compute Services',
        'Storage',
        'Security Services',
        'DB Services'
    ]
    
    for category in categories:
        column = f"{category} %"
        avg, max_val, min_val, max_account, min_account = get_stats(column)
        
        summary_data.extend([
            {'Metric': f'{category} % Statistics', 'Value': ''},
            {'Metric': 'Industry Average', 'Value': f"{avg:.2f}%"},
            {'Metric': 'Highest Usage', 'Value': f"{max_val:.2f}% ({max_account})"},
            {'Metric': 'Lowest Usage', 'Value': f"{min_val:.2f}% ({min_account})"}
        ])
        summary_data.append({})  # Empty row
    
    # Add extra empty row before detailed data
    summary_data.append({})
    
    return pd.DataFrame(summary_data)


def create_service_category_analysis(cleaned_data):
    """Create analysis of service category percentages based on domain header rows"""
    service_categories = get_service_domains()
    analysis_data = []
    
    for account_name, df in cleaned_data.items():
        try:
            total_spend = get_account_total_spend(df)
            print(f"Processing {account_name} with total spend: ${total_spend:,.2f}")
            
            category_spend = {category: 0 for category in service_categories.keys()}
            uncategorized_spend = 0
            
            # First pass to calculate categorized spend
            for idx, row in df.iterrows():
                service_name = str(row.iloc[0])
                if pd.isna(service_name) or 'View AWS service' in service_name:
                    continue
                
                spend = clean_amount(row.iloc[1])
                categorized = False
                
                for domain in service_categories.keys():
                    if service_name == domain:
                        category_spend[domain] = spend
                        categorized = True
                        break
                
                if not categorized:
                    uncategorized_spend += spend
            
            # Find highest spend domain
            if category_spend:
                highest_domain = max(category_spend.items(), key=lambda x: x[1])
                highest_domain_name = highest_domain[0]
                highest_domain_spend = highest_domain[1]
                highest_domain_pct = (highest_domain_spend / total_spend * 100) if total_spend > 0 else 0
            else:
                highest_domain_name = "None"
                highest_domain_spend = 0
                highest_domain_pct = 0
            
            row_data = {
                'Customer Name': account_name,
                'Total Spend': total_spend,
                'Highest Spend Domain': f"{highest_domain_name} (${highest_domain_spend:,.2f}, {highest_domain_pct:.2f}%)"
            }
            
            # Calculate percentages for service categories only (excluding uncategorized)
            percentages_sum = 0
            for category in service_categories.keys():
                spend = category_spend[category]
                percentage = (spend / total_spend * 100) if total_spend > 0 else 0
                row_data[f"{category} %"] = percentage
                percentages_sum += percentage
            
            # Add sum of service category percentages
            row_data['Total Category %'] = percentages_sum
            
            analysis_data.append(row_data)
            
        except Exception as e:
            print(f"Error processing {account_name}: {str(e)}")
            continue

    # Create DataFrame
    df_analysis = pd.DataFrame(analysis_data)
    
    # Calculate averages before adding the average row
    spend_avg = df_analysis['Total Spend'].mean()
    category_avgs = {}
    for category in service_categories.keys():
        category_avgs[f"{category} %"] = df_analysis[f"{category} %"].mean()
    
    # Add industry average row
    avg_row = {
        'Customer Name': 'Industry Average',
        'Total Spend': spend_avg,
        'Highest Spend Domain': 'Average across all accounts'
    }
    avg_row.update({k: v for k, v in category_avgs.items()})
    avg_row['Total Category %'] = sum(category_avgs.values())
    
    df_analysis = pd.concat([df_analysis, pd.DataFrame([avg_row])], ignore_index=True)
    
    # Format all percentage columns
    percentage_columns = [col for col in df_analysis.columns if col.endswith(' %')]
    for col in percentage_columns:
        df_analysis[col] = df_analysis[col].apply(
            lambda x: f"{x:.2f}%"
        )
    
    # Format total spend
    df_analysis['Total Spend'] = df_analysis['Total Spend'].apply(
        lambda x: f"${x:,.2f}"
    )
    
    # Add top categories
    df_analysis['Top Categories'] = df_analysis.apply(
        lambda row: get_top_categories(row, service_categories.keys()), 
        axis=1
    )
    
    return df_analysis



def save_analysis_to_excel(output_file, df_summary, df_monthly_trends, 
                         df_cumulative, cleaned_data):
    """Save all analysis results to Excel file with combined analysis"""
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Create and save overall summary
            print("Creating Overall Summary...")
            df_overall_summary = create_overall_summary(cleaned_data, account_domain_spend)
            df_overall_summary.to_excel(writer, sheet_name='Overall Summary', index=False)
            
            # Create service group analysis
            print("Creating Service Group Analysis...")
            df_service_groups = create_service_category_analysis(cleaned_data)
            
            # Create specialized service summaries
            print("Creating Service-Specific Summaries...")
            df_ml_summary = create_ml_services_summary(account_domain_spend)
            df_compute_summary = create_compute_services_summary(account_domain_spend)
            df_storage_summary = create_storage_services_summary(account_domain_spend)
            df_database_summary = create_database_services_summary(account_domain_spend)
            
            # Save all sheets in desired order
            print("Saving all analysis sheets...")
            
            # 1. Overall Summary (already saved)
            # 2. Service Group Analysis
            df_service_groups.to_excel(writer, sheet_name='Service Group Analysis', index=False)
            
            # 3. Service Summaries
            df_ml_summary.to_excel(writer, sheet_name='ML Services Summary', index=False)
            df_compute_summary.to_excel(writer, sheet_name='Compute Services Summary', index=False)
            df_storage_summary.to_excel(writer, sheet_name='Storage Services Summary', index=False)
            df_database_summary.to_excel(writer, sheet_name='Database Services Summary', index=False)
            
            # 4. Other Analysis Sheets
            df_summary.to_excel(writer, sheet_name='Domain Summary', index=False)
            df_monthly_trends.to_excel(writer, sheet_name='Detailed Analysis', index=False)
            
            # Get workbook and format all sheets
            workbook = writer.book
            
            print("Applying formatting to sheets...")
            
            # Format Overall Summary
            worksheet = writer.sheets['Overall Summary']
            format_sheet(worksheet, workbook, is_summary=True)
            
            # Format Service Group Analysis
            worksheet = writer.sheets['Service Group Analysis']
            format_excel_with_highlights(worksheet, 'Service Group Analysis')
            
            # Format Service Summaries
            service_summary_sheets = [
                'ML Services Summary',
                'Compute Services Summary',
                'Storage Services Summary',
                'Database Services Summary'
            ]
            
            for sheet_name in service_summary_sheets:
                print(f"Formatting {sheet_name}...")
                worksheet = writer.sheets[sheet_name]
                format_service_summary_sheet(worksheet, sheet_name)
            
            # Format remaining Analysis Sheets
            analysis_sheets = [
                'Domain Summary',
                'Detailed Analysis'
            ]
            
            for sheet_name in analysis_sheets:
                print(f"Formatting {sheet_name}...")
                worksheet = writer.sheets[sheet_name]
                format_excel_with_highlights(worksheet, sheet_name)
            
            # Adjust column widths for all sheets
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                adjust_column_widths(worksheet)
            
            print("Adding final touches...")
            # Add filters and freeze panes for all sheets
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                worksheet.freeze_panes = 'A2'
                worksheet.auto_filter.ref = worksheet.dimensions
            
            print(f"Successfully saved analysis to {output_file}")
            
    except Exception as e:
        print(f"Error saving Excel file: {str(e)}")
        raise


def format_service_summary_sheet(worksheet, sheet_name):
    """Apply specialized formatting for service summary sheets"""
    # Define styles
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    summary_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
    above_avg_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green
    white_font = Font(color='FFFFFF', bold=True)
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Get the percentage column index
    percentage_col = None
    for idx, cell in enumerate(worksheet[1], 1):
        if 'Services %' in str(cell.value):
            percentage_col = idx
            break

    # If we found the percentage column, calculate average
    if percentage_col:
        # Skip summary rows (first 7 rows) and get all percentage values
        percentages = []
        for row in range(9, worksheet.max_row + 1):  # Start after summary section
            cell_value = worksheet.cell(row=row, column=percentage_col).value
            if isinstance(cell_value, str) and '%' in cell_value:
                try:
                    percentage = float(cell_value.replace('%', ''))
                    percentages.append(percentage)
                except ValueError:
                    continue

        if percentages:
            avg_percentage = sum(percentages) / len(percentages)

            # Now highlight cells above average
            for row in range(9, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=percentage_col)
                if isinstance(cell.value, str) and '%' in cell.value:
                    try:
                        value = float(cell.value.replace('%', ''))
                        if value > avg_percentage:
                            cell.fill = above_avg_fill
                    except ValueError:
                        continue

    # Format summary section
    for row in range(1, 8):  # Summary section rows
        for cell in worksheet[row]:
            if row == 1:  # Title row
                cell.font = bold_font
                cell.fill = summary_fill
            cell.alignment = center_align
    
    # Format column headers for detailed section
    for cell in worksheet[8]:  # Header row for detailed section
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = center_align
    
    # Format data section
    for row in range(9, worksheet.max_row + 1):
        for cell in worksheet[row]:
            cell.alignment = center_align
            
            # Right align amounts and percentages
            if isinstance(cell.value, str):
                if cell.value.startswith('$') or cell.value.endswith('%'):
                    cell.alignment = right_align
    
    # Add conditional formatting for percentage columns
    if sheet_name == 'Compute Services Summary':
        for col_letter in ['D', 'E']:  # Both percentage columns
            start_row = 9  # After summary section
            data_range = f"{col_letter}{start_row}:{col_letter}{worksheet.max_row}"
            
            # Get values for the column to calculate average
            values = []
            for row in range(start_row, worksheet.max_row + 1):
                cell = worksheet[f"{col_letter}{row}"]
                if isinstance(cell.value, str) and '%' in cell.value:
                    try:
                        value = float(cell.value.replace('%', ''))
                        values.append(value)
                    except ValueError:
                        continue

            if values:
                avg_value = sum(values) / len(values)
                
                # Highlight cells above average
                for row in range(start_row, worksheet.max_row + 1):
                    cell = worksheet[f"{col_letter}{row}"]
                    if isinstance(cell.value, str) and '%' in cell.value:
                        try:
                            value = float(cell.value.replace('%', ''))
                            if value > avg_value:
                                cell.fill = above_avg_fill
                        except ValueError:
                            continue
    
    # Adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def adjust_column_widths(worksheet):
    """Adjust column widths based on content"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # Set width with some padding, but not too wide
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width




def adjust_column_widths(worksheet):
    """Adjust column widths based on content"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # Set width with some padding, but not too wide
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width



if __name__ == "__main__":
    try:
        excel_file_path = 'MyPodData.xlsx'
        
        if not os.path.exists(excel_file_path):
            raise FileNotFoundError(f"The file {excel_file_path} does not exist")
        
        print("Starting AWS spend analysis...")
        print("=" * 50)
        
        # Load and clean data
        print("\nStep 1: Loading and cleaning data...")
        cleaned_data = clean_excel_data(excel_file_path)
        
        if not cleaned_data:
            raise ValueError("No valid data found in Excel file")
        
        # Perform analyses
        print("\nStep 2: Performing comprehensive analysis...")
        
        # Domain Summary Analysis
        print("Processing Domain Summary...")
        df_summary, account_domain_spend = analyze_service_domains(cleaned_data)
        
        # Comparative Analysis
        print("Creating Comparative Analysis...")
        df_comparative, df_trends = create_comparative_analysis(cleaned_data, account_domain_spend)
        
        # Monthly Patterns Analysis
        print("Analyzing Monthly Patterns...")
        monthly_patterns = analyze_monthly_patterns(cleaned_data, account_domain_spend)
        df_monthly_trends = create_monthly_trends_report(monthly_patterns)
        
        # Cumulative Analysis
        print("Generating Cumulative Analysis...")
        df_cumulative = create_cumulative_report(cleaned_data, account_domain_spend)
        
        # Save to Excel
        output_file = 'AWS_Complete_Spend_Analysis.xlsx'
        print(f"\nStep 3: Saving analysis to {output_file}...")
        save_analysis_to_excel(
            output_file,
            df_summary,
            df_monthly_trends,
            df_cumulative,
            cleaned_data
        )
        
        print("\nAnalysis Complete!")
        print("=" * 50)
        print("\nOutput Excel file contains:")
        print("1. Service Categories:")
        print("   - Detailed breakdown of service usage percentages")
        print("   - Industry average comparisons")
        print("   - Service concentration analysis")
        
        print("\n2. Domain Summary:")
        print("   - Overall service domain spending overview")
        print("   - Domain-wise total and average spend")
        print("   - Account distribution across domains")
        
        print("\n3. Account Comparison:")
        print("   - Account-level spending patterns")
        print("   - Domain concentration analysis")
        print("   - Month-over-month changes")
        
        print("\n4. Monthly Trends:")
        print("   - Monthly spending patterns")
        print("   - Seasonal variations")
        print("   - Growth trends")
        
        print("\n5. Detailed Analysis:")
        print("   - Comprehensive spending analysis")
        print("   - Anomaly detection")
        print("   - Risk assessment")
        
        print("\n6. Cumulative Analysis:")
        print("   - Cross-domain comparisons")
        print("   - Account spending distributions")
        print("   - Domain utilization patterns")
        
        # Print key findings
        print("\nKey Findings:")
        print("-" * 30)
        
        # Service Category Insights
        df_service_categories = create_service_category_analysis(cleaned_data)
        print("\nTop Service Categories:")
        for _, row in df_service_categories.iterrows():
            if row['Customer Name'] == 'Industry Average':
                for col in df_service_categories.columns:
                    if col.endswith('%'):
                        value = float(row[col].replace('%', ''))
                        if value > 10:  # Show categories with >10% usage
                            print(f"- {col.replace(' %', '')}: {row[col]}")
        
        # High risk patterns
        high_risk_patterns = df_monthly_trends[df_monthly_trends['Risk Level'] == 'High']
        if not high_risk_patterns.empty:
            print("\nHigh Risk Spending Patterns Detected:")
            for _, pattern in high_risk_patterns.iterrows():
                print(f"- {pattern['Account']} - {pattern['Service Domain']}")
                print(f"  * Pattern: {pattern['Spending Pattern']}")
                print(f"  * Changes: {pattern['Notable Changes']}")
        
        # Top spending domains
        top_domains = df_summary.head(3)
        print("\nTop 3 Spending Domains:")
        for _, domain in top_domains.iterrows():
            print(f"- {domain['Service Domain']}: {domain['Total Spend']}")
        
        # Significant variations
        significant_variations = df_monthly_trends[
            df_monthly_trends['Volatility'].str.rstrip('%').astype(float) > 50
        ]
        if not significant_variations.empty:
            print("\nSignificant Spending Variations:")
            for _, var in significant_variations.iterrows():
                print(f"- {var['Account']} - {var['Service Domain']}")
                print(f"  * Volatility: {var['Volatility']}")
        
        print("\nRecommended Actions:")
        print("1. Review high-risk spending patterns")
        print("2. Investigate significant variations")
        print("3. Optimize top spending domains")
        print("4. Monitor monthly trends for cost optimization")
        print("5. Review service category distribution")
        
        print("\nAnalysis file saved successfully!")
        print(f"Location: {os.path.abspath(output_file)}")
        
    except Exception as e:
        print(f"\nError occurred: {str(e)}")
        print("\nStack trace:")
        import traceback
        traceback.print_exc()
        raise
